import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QPushButton, QLabel, QFileDialog, QDialog, QDateTimeEdit, QMessageBox, QProgressDialog, QTableView, QButtonGroup, QRadioButton, QLineEdit)
from PyQt5.QtCore import QDateTime, Qt, QThread, pyqtSignal
from sql_setting_dialog import ConfigDialog
from config import ConfigManager
from datetime import datetime, time
from orcale_db_connect import connect_to_oracle_test, get_his_data
from sqlserver_db_connect import connect_to_sqlserver_test, import_data_to_yy
from log_util import logger
from path_util import open_with_default_app
import traceback
import sys
from scroll_window import SqlInfiniteTableWidget
from py_sqlite import SQLiteHelper
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta
from checkable_combo_box import CheckableComboBox
from yy_dept_mapper import get_shoukuan
from his_data_dialog import PaymentRecordDialog

format_pattern = "yyyy-MM-dd HH:mm:ss"
class ExcelMerger(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("用友导入工具")

        # 1. 设置窗口的大小
        width, height = 800, 500

        # 2. 获取屏幕对象和屏幕几何信息
        screen = QApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()  # 获取可用区域（排除任务栏）

        # 3. 计算居中坐标
        # x = (屏幕宽 - 窗口宽) / 2
        # y = (屏幕高 - 窗口高) / 2
        x = (screen_geometry.width() - width) // 2
        y = (screen_geometry.height() - height) // 2

        self.setGeometry(x, y, width, height)
        
        # 创建主窗口部件
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        # 创建垂直布局
        layout = QVBoxLayout()
        main_widget.setLayout(layout)

        self.config_manager = ConfigManager()
        # 数据库设置
        template_layout = QHBoxLayout()
        # 数据库设置
        template_layout.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.template_sql_setting = QPushButton("系统设置")
        self.template_sql_setting.setFixedSize(100, 30)
        self.template_sql_setting.clicked.connect(self.show_config_dialog)
        template_layout.addWidget(self.template_sql_setting)
        template_layout.addStretch(1)
        layout.addLayout(template_layout)

        his_layout = QHBoxLayout()
        self.his_label = QLabel("his配置:")
        self.his_test_connect = QPushButton("测试连接")
        self.his_test_connect.setFixedSize(80, 20)
        self.his_test_connect.clicked.connect(self.his_test_connect_click)
        his_layout.addWidget(self.his_label)
        his_layout.addWidget(self.his_test_connect)

        yy_layout = QHBoxLayout()
        self.sql_label = QLabel("用友配置:")
        self.sql_test_connect = QPushButton("测试连接")
        self.sql_test_connect.setFixedSize(80, 20)
        self.sql_test_connect.clicked.connect(self.yy_test_connect_click)
        yy_layout.addWidget(self.sql_label)
        yy_layout.addWidget(self.sql_test_connect)
        self.update_config_text()

        layout.addLayout(his_layout)
        layout.addLayout(yy_layout)

        # 业务类型
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("业务类型："))

        # 创建单选按钮
        self.radio_outpatient = QRadioButton("门诊收入")
        self.radio_inpatient1 = QRadioButton("住院结算")
        self.radio_inpatient2 = QRadioButton("全院病人费用(月)")
        self.radio_inpatient3 = QRadioButton("门诊自助机(月)")
        self.radio_inpatient4 = QRadioButton("门诊扫码(月)")

        self.radio_outpatient.setChecked(True)

        # 使用 QButtonGroup 进行逻辑分组（确保互斥）
        self.type_group = QButtonGroup(self)
        self.type_group.addButton(self.radio_outpatient, 0)
        self.type_group.addButton(self.radio_inpatient1, 1)
        self.type_group.addButton(self.radio_inpatient2, 2)
        self.type_group.addButton(self.radio_inpatient3, 3)
        self.type_group.addButton(self.radio_inpatient4, 4)
        self.type_group.idClicked.connect(self.on_type_changed)
        # 添加到布局
        type_layout.addWidget(self.radio_outpatient)
        type_layout.addWidget(self.radio_inpatient1)
        type_layout.addWidget(self.radio_inpatient2)
        type_layout.addWidget(self.radio_inpatient3)
        type_layout.addWidget(self.radio_inpatient4)
        type_layout.addStretch()  # 添加弹簧，让按钮靠左对齐
        layout.addLayout(type_layout)

        # 日导出初始日期选择为昨天
        prev_day = get_prev_day()
        self.start_str = f"{prev_day} 00:00:00"
        self.end_str = f"{prev_day} 23:59:59"

        # 时间选择
        time_layout = QHBoxLayout()
        # 创建日期时间选择器
        self.dt_edit = QDateTimeEdit(self)
        self.dt_edit.setDateTime(QDateTime.fromString(self.start_str, format_pattern))
        self.dt_edit.setDisplayFormat(format_pattern)
        self.dt_edit.setCalendarPopup(True)
        self.dt_edit.dateTimeChanged.connect(lambda dt: self.on_date_changed(dt, 0))
        time_layout.addWidget(QLabel("扎帐开始时间："))
        time_layout.addWidget(self.dt_edit)
        layout.addLayout(time_layout)
        time_layout1 = QHBoxLayout()
        # 创建日期时间选择器
        self.dt_edit1 = QDateTimeEdit(self)
        self.dt_edit1.setDateTime(QDateTime.fromString(self.end_str, format_pattern))
        self.dt_edit1.setDisplayFormat(format_pattern)
        self.dt_edit1.setCalendarPopup(True)
        self.dt_edit1.dateTimeChanged.connect(lambda dt: self.on_date_changed(dt, 1))
        time_layout1.addWidget(QLabel("扎帐结束时间："))
        time_layout1.addWidget(self.dt_edit1)
        layout.addLayout(time_layout1)

        self.shouyin_layout = QHBoxLayout()
        self.shouyin_box = CheckableComboBox()
        self.shouyin_box.add_all_option()  # 先加全选项
        self.shouyin_box.addItem(get_shoukuan())
        self.shouyin_layout.addWidget(QLabel("收银员"))
        self.shouyin_layout.addWidget(self.shouyin_box)
        layout.addLayout(self.shouyin_layout)

        self.code_layout = QHBoxLayout()
        self.code_edit = QLineEdit()
        self.code_edit.setPlaceholderText("请输入扎帐单号，不输入则不筛选")
        self.code_edit.setFixedWidth(390)
        self.code_layout.addWidget(QLabel("扎帐单号:"))
        self.code_layout.addStretch()
        self.code_layout.addWidget(self.code_edit)
        layout.addLayout(self.code_layout)

        # 开始按钮
        self.full_path = ""
        self.full_path_num = 0
        start_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始查询")
        self.start_btn.setEnabled(True)
        self.start_btn.clicked.connect(self.start_his_export)
        start_layout.addWidget(self.start_btn)

        layout.addLayout(start_layout)

        self.table = SqlInfiniteTableWidget(main_view=self)
        self.table.setFixedHeight(400)
        layout_sql = QVBoxLayout()
        layout_sql.addWidget(QLabel("操作记录"))
        layout_sql.addWidget(self.table)
        layout.addLayout(layout_sql)
        layout.addStretch(1)

        self.sqlite_helper = SQLiteHelper()
        self.record_id = -1
        

    def show_config_dialog(self):
        # 实例化弹窗
        dialog = ConfigDialog(self.config_manager, parent=self)
        # 运行弹窗 阻塞主窗口，直到弹窗关闭
        if dialog.exec_() == QDialog.Accepted:
            # 如果点击了“确定”，获取数据并更新主界面
            self.update_config_text()
        else:
            print("用户取消了输入")

    def show_his_data_dialog(self, start_str, end_str, type, shoukuan_df, total_df):
        # 实例化弹窗
        dialog = PaymentRecordDialog(start_str, end_str, type, shoukuan_df, total_df, self.config_manager, parent=self)
        # 运行弹窗 阻塞主窗口，直到弹窗关闭
        if dialog.exec_() == QDialog.Accepted:
            # 如果点击了“确定”，获取数据并更新主界面
            self.update_config_text()
        else:
            print("用户取消了输入")

    def update_config_text(self):
        """ 更新配置显示 """
        print("update_config_text")
        oracle_config = self.config_manager.get_db_config("oracle")
        oracle_config_text = f"his配置: {oracle_config.get('ip')}:{oracle_config.get('port')}/{oracle_config.get('service_name')} user:{oracle_config.get('user')}  password:{oracle_config.get('password')}"
        self.his_label.setText(oracle_config_text)
        sqlserver_config = self.config_manager.get_db_config("sqlserver")
        sqlserver_config_text = f"用友配置: {sqlserver_config.get('ip')}:{sqlserver_config.get('port')}/{sqlserver_config.get('database')} user:{sqlserver_config.get('user')}  password:{sqlserver_config.get('password')}"
        self.sql_label.setText(sqlserver_config_text)

    def on_date_changed(self, q_dt, type):
        """ 时间选择变化 """
        if type == 0:
            self.start_str = q_dt.toString(format_pattern)
        else:
            self.end_str = q_dt.toString(format_pattern)

    def on_type_changed(self, type_id):
        now = datetime.now()
        yesterday = get_prev_day()
        if type_id >= 2:
            # 门诊扫码、自助机、住院病人时间选择切换到月
            # 减去 1 个月
            last_month_today = now - relativedelta(months=1)
            self.start_str = f"{last_month_today.strftime('%Y-%m-%d')} 00:00:00"
            # 禁用
            self.shouyin_box.setEnabled(False)
            self.code_edit.setEnabled(False)
        else:
            self.start_str = f"{yesterday} 00:00:00"
            self.shouyin_box.setEnabled(True)
            self.code_edit.setEnabled(True)
        self.dt_edit.setDateTime(QDateTime.fromString(self.start_str, format_pattern))
        self.end_str = f"{yesterday} 23:59:59"

    def his_test_connect_click(self):
        """ 测试his连接 """
        oracle_config = self.config_manager.get_db_config("oracle")
        result = connect_to_oracle_test(oracle_config.get('user'), oracle_config.get('password'), oracle_config.get('ip'),
                               oracle_config.get('port'), oracle_config.get('service_name'))
        if result:
            QMessageBox.information(self, "成功", "his系统测试连接成功！")
        else:
            QMessageBox.warning(self, "错误", "his系统测试连接失败！")

    def yy_test_connect_click(self):
        """ 测试用友连接 """
        sqlserver_config = self.config_manager.get_db_config("sqlserver")
        result = connect_to_sqlserver_test(sqlserver_config.get('ip'), sqlserver_config.get('port'),
                                        sqlserver_config.get('database'),
                                        sqlserver_config.get('user'), sqlserver_config.get('password'))
        if result:
            QMessageBox.information(self, "成功", "用友系统测试连接成功！")
        else:
            QMessageBox.warning(self, "错误", "用友系统测试连接失败！")

    def start_his_export(self):
        """ 开始查询His """
        if not self.start_str or not self.end_str:
            QMessageBox.warning(self, "错误", "请选择导出的时间")
            logger.info(f"开始查询His: 请选择导出的时间")
            return
        try:
            # pd.to_datetime 会自动解析格式，.date() 会去掉时分秒
            is_same_day = pd.to_datetime(self.start_str).date() == pd.to_datetime(self.end_str).date()
        except Exception:
            QMessageBox.warning(self, "错误", "请选择导出的时间")
            logger.info(f"开始查询His: 请选择导出的时间")
            return

        if not is_same_day:
            QMessageBox.warning(self, "错误", "时间跨度应该为一天")
            logger.info(f"开始查询His: 时间跨度应该为一天")
            return

        if self.start_str and self.end_str:
            logger.info(f"开始查询His: {self.start_str}, {self.end_str}")
            progress = QProgressDialog("系统正在处理中，请稍候...", None, 0, 0)
            progress.setWindowTitle("请等待")
            progress.setWindowModality(Qt.WindowModal)
            progress.setCancelButton(None)
            progress.show()

            # 强制刷新界面渲染加载窗
            QApplication.processEvents()
            # 执行查询
            # shoukuan_df, total_df, full_path = get_his_data(self.config_manager.get_db_config('oracle'), self.start_str, self.end_str, self.type_group.checkedId(), self.shouyin_box.checked_items(), self.code_edit.text().strip())
            full_path = r"C:\Users\Administrator\Desktop\线上his\门诊收入2026-02-04 00_00_00数据导出20_56_28.xlsx"
            df = pd.read_excel(full_path, sheet_name=['收款数据', '汇总数据'])
            shoukuan_df = df['收款数据']
            total_df = df['汇总数据']
            # 关闭加载窗
            progress.close()
            logger.info(f"导出地址：{full_path}")
            self.show_his_data_dialog(self.start_str, self.end_str, self.type_group.checkedId(), shoukuan_df, total_df)


    def start_yy_import(self, item=None):
        """ 导入用友数据 """
        if item:
            this_record_id = item['id']
            this_full_path = item['export_file_path']
        else:
            this_record_id = self.record_id
            this_full_path = self.full_path

        # this_full_path = r"C:\Users\Administrator\Desktop\线上his\门诊2026-01-29 00_00_00数据导出.xlsx"
        if this_full_path and this_record_id > 0:
            # 参数说明：父窗口, 标题, 内容, 按钮组合, 默认选中的按钮
            record = self.sqlite_helper.get_record_by_id(this_record_id)
            # record['data_type'] = '0'
            # record['export_file_path'] = r"C:\Users\Administrator\Desktop\线上his\门诊2026-01-29 00_00_00数据导出.xlsx"
            # record['data_type'] = '1'
            # record['export_file_path'] = r"C:\Users\Administrator\Desktop\线上his\住院缴费2026-01-29 00_00_00数据导出1.xlsx"
            # record['data_type'] = '2'
            # record['export_file_path'] = r"C:\Users\Administrator\Desktop\线上his\全院病人费用2026-01-05 00_00_00数据导出20_46_15.xlsx"
            # record['data_type'] = '3'
            # record['export_file_path'] = r"C:\Users\Administrator\Desktop\线上his\门诊自助机2026-02-04 00_00_00数据导出20_58_20.xlsx"
            # record['data_type'] = '4'
            # record['export_file_path'] = r"C:\Users\Administrator\Desktop\线上his\门诊扫码2026-02-04 00_00_00数据导出20_56_48.xlsx"

        else:
            QMessageBox.warning(self, "失败", "请点击导出HIS数据后进行导入")

    def _start_import_data_to_yy(self, record):
        """
        公共用友导入方法
        """
        logger.info(f"开始用友导入: {record}")
        def on_complete(success, message, record):
            # 更新本地数据库状态
            if success:
                logger.info("导入成功，刷新列表")
                self.sqlite_helper.import_yy_result(record["id"], success, message, record["import_yy_num"])
                QMessageBox.information(self, "成功", "导入数据成功")
            else:
                logger.info("导入失败，刷新列表")
                QMessageBox.critical(self, "失败", f"导入失败{message}")
                self.sqlite_helper.import_yy_result(record["id"], success, "", -1)
            self.table.refresh_data()


    def run_with_loading(parent, task_func):
        """
        通用加载窗工具
        :param parent: 父窗口 (self)
        :param task_func: 需要执行的耗时函数名 (不要带括号)
        :param args: 传递给 task_func 的位置参数
        :param kwargs: 传递给 task_func 的关键字参数
        """
        # 1. 创建并显示进度条
        progress = QProgressDialog("系统正在处理中，请稍候...", None, 0, 0, parent)
        progress.setWindowTitle("请等待")
        progress.setWindowModality(Qt.WindowModal)
        progress.setCancelButton(None)
        progress.show()

        # 强制刷新界面渲染加载窗
        QApplication.processEvents()

        result = None
        try:
            # 2. 执行传入的方法
            # 如果 task_func 是 df.to_excel，则这里实际执行 df.to_excel(*args, **kwargs)
            task_func()
        except Exception as e:
            print(f"执行出错: {traceback.format_exc()}")
            QMessageBox.critical(parent, "错误", f"操作失败: {str(e)}")
        finally:
            # 3. 关闭加载窗
            progress.close()


def merge_excel_files(template_path, folder_path, output_path):
    # 读取模板文件以获取结构和公式
    template_wb = load_workbook(template_path)
    template_ws = template_wb.active

    # 获取所有excel文件
    excel_files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls'))]

    # 用于存储所有数据框的列表
    all_dfs = []
    file_names = []

    # 读取每个Excel文件
    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        try:
            df = pd.read_excel(file_path)
            all_dfs.append(df)
            file_names.append(os.path.splitext(file)[0])  # 获取文件名（不含扩展名）
        except Exception as e:
            print(f"无法读取文件 {file}: {str(e)}")

    if not all_dfs:
        print("没有找到可以处理的Excel文件")
        return

    # 合并所有数据框
    merged_df = pd.concat(all_dfs, ignore_index=True)

    # 根据所有列进行分组并求和
    numeric_columns = merged_df.select_dtypes(include=['int64', 'float64']).columns
    if len(numeric_columns) > 0:
        group_columns = [col for col in merged_df.columns if col not in numeric_columns]
        if group_columns:
            result_df = merged_df.groupby(group_columns, as_index=False)[numeric_columns].sum()
        else:
            result_df = merged_df
    else:
        result_df = merged_df

    # 创建新的工作簿
    output_wb = load_workbook(template_path)
    output_ws = output_wb.active

    # 清空现有数据（保留标题行）
    for row in range(2, output_ws.max_row + 1):
        for col in range(1, output_ws.max_column + 1):
            output_ws.cell(row=row, column=col, value=None)

    # 将求和数据写入第一个sheet页，保留公式
    for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            # 获取模板中对应单元格的公式
            template_cell = template_ws.cell(row=r_idx, column=c_idx)
            if template_cell.value and str(template_cell.value).startswith('='):
                # 如果是公式，保留公式
                output_ws.cell(row=r_idx, column=c_idx, value=template_cell.value)
            else:
                # 否则写入数据
                output_ws.cell(row=r_idx, column=c_idx, value=value)

    # 将第一个sheet页重命名为"汇总"
    output_ws.title = "汇总"

    # 处理每个Excel文件，创建额外的sheet页
    for df, sheet_name in zip(all_dfs, file_names):
        # 创建新的sheet页
        if sheet_name in output_wb.sheetnames:
            # 如果sheet名已存在，添加序号
            i = 1
            while f"{sheet_name}_{i}" in output_wb.sheetnames:
                i += 1
            sheet_name = f"{sheet_name}_{i}"

        # 创建新的工作表
        new_ws = output_wb.create_sheet(title=sheet_name)

        # 复制模板工作表的格式和公式
        for row in template_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                if cell.value and str(cell.value).startswith('='):
                    # 如果是公式，保留公式
                    new_cell.value = cell.value
                elif cell.row == 1:
                    # 如果是标题行，保留标题
                    new_cell.value = cell.value

        # 将数据写入工作表
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                # 获取模板中对应单元格的公式
                template_cell = template_ws.cell(row=r_idx, column=c_idx)
                if not (template_cell.value and str(template_cell.value).startswith('=')):
                    # 如果不是公式，写入数据
                    new_ws.cell(row=r_idx, column=c_idx, value=value)

    # 保存结果
    output_wb.save(output_path)
    print(f"合并完成，结果已保存至: {output_path}")

def get_prev_day():
    yesterday = datetime.now() - relativedelta(days=1)
    return yesterday.strftime('%Y-%m-%d')


def exception_hook(exctype, value, tb):
    """全局异常捕获钩子"""
    err_msg = "".join(traceback.format_exception(exctype, value, tb))
    print(err_msg)  # 打印到终端
    # 也可以弹窗显示，防止闪退后看不到
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("程序崩溃了！")
    msg.setInformativeText(err_msg)
    msg.setWindowTitle("Fatal Error")
    msg.exec_()
    sys.exit(1)

if __name__ == "__main__":
    # 必须在实例化 QApplication 之前调用
    # 将系统默认的异常处理器替换为我们自定义的
    sys.excepthook = exception_hook
    QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    app = QApplication([])
    window = ExcelMerger()
    window.show()
    app.exec_()
